import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import matplotlib.pyplot as plt
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image

st.set_page_config(layout="wide", page_title="Análisis de Fusión - Eléctricos Internacional")

# Custom CSS for branding
st.markdown("""
<style>
    /* Main color palette based on #002f6c */
    :root {
        --primary-blue: #002f6c;
        --secondary-blue: #004a9c;
        --light-blue: #e8f0f8;
        --accent-orange: #ff6b35;
        --light-gray: #f8f9fa;
        --dark-gray: #343a40;
    }
    
    /* Header styling */
    .main-header {
        background: linear-gradient(135deg, #002f6c 0%, #004a9c 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    
    .main-header h1 {
        color: white;
        margin: 0;
        font-size: 2.5rem;
        font-weight: 700;
    }
    
    .main-header p {
        color: #e8f0f8;
        margin: 0.5rem 0 0 0;
        font-size: 1.1rem;
    }
    
    /* Section headers */
    h2, h3 {
        color: #002f6c !important;
        padding-bottom: 0.5rem;
        margin-top: 2rem !important;
    }
    
    /* Metrics styling */
    [data-testid="stMetricValue"] {
        color: #002f6c;
        font-size: 1.8rem !important;
        font-weight: 700;
    }
    
    [data-testid="stMetricLabel"] {
        color: #004a9c;
        font-weight: 600;
    }
    
    /* Button styling */
    .stDownloadButton button {
        background-color: #002f6c !important;
        color: white !important;
        border: none !important;
        padding: 0.75rem 2rem !important;
        font-size: 1.1rem !important;
        font-weight: 600 !important;
        border-radius: 8px !important;
        transition: all 0.3s ease !important;
    }
    
    .stDownloadButton button:hover {
        background-color: #004a9c !important;
        box-shadow: 0 4px 12px rgba(0, 47, 108, 0.3) !important;
        transform: translateY(-2px) !important;
    }
    
    /* Slider styling */
    .stSlider [data-baseweb="slider"] {
        background-color: #e8f0f8;
    }
    
    .stSlider [data-baseweb="slider"] [role="slider"] {
        background-color: #002f6c !important;
    }
    
    /* Info boxes */
    .stAlert {
        background-color: #e8f0f8;
        border-left: 4px solid #002f6c;
    }
    
    /* Dataframe styling */
    [data-testid="stDataFrame"] {
        border: 2px solid #e8f0f8;
        border-radius: 8px;
    }
    
    /* File uploader */
    [data-testid="stFileUploader"] {
        background-color: #f8f9fa;
        border: 2px dashed #004a9c;
        border-radius: 8px;
        padding: 1rem;
    }
    
    /* Divider */
    hr {
        border-color: #e8f0f8 !important;
        border-width: 2px !important;
    }
    
    /* Logo container */
    .logo-container {
        display: flex;
        justify-content: center;
        align-items: center;
        padding: 1rem 0;
        margin-bottom: 1rem;
    }
    
    .logo-container img {
        max-width: 400px;
        height: auto;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="logo-container">
    <img src="https://hebbkx1anhila5yf.public.blob.vercel-storage.com/Logo-MxBdLC0RPJjso0OMZnL9V1i3uW5Fln.png" alt="Eléctricos Internacional SAS">
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>Herramienta de Análisis de Fusión de Fusibles</h1>
    <p>Sube tu archivo CSV para identificar automáticamente las fases de transitorio inicial, estado estable y fusión.</p>
</div>
""", unsafe_allow_html=True)

# --- Nueva Función para Crear el Gráfico Estático (para Excel) ---
def create_matplotlib_figure(df, results, threshold):
    """Crea una figura estática con Matplotlib para el reporte."""
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.plot(df['time_s'], df['current_kA'], label='Señal Original i(t)', alpha=0.7, color='royalblue')
    ax.plot(df['time_s'], df['i_rms'], label='RMS Móvil i_rms(t)', color='red')
    
    ax.axhline(y=threshold, color='green', linestyle='--', label='Umbral')
    
    # Extraer valores de los resultados para las líneas verticales
    t_inicio_evento = results.loc[results['Parámetro'] == 'Tiempo de Inicio del Evento (s)', 'Valor'].iloc[0]
    t_pico_energia = results.loc[results['Parámetro'] == 'Tiempo del Pico de Energía (s)', 'Valor'].iloc[0]
    t_inicio_fusion = results.loc[results['Parámetro'] == 'Tiempo de Inicio de Fusión (s)', 'Valor'].iloc[0]
    t_fin_evento = results.loc[results['Parámetro'] == 'Tiempo de Fin del Evento (s)', 'Valor'].iloc[0]

    # Se convierte a float para asegurar que Matplotlib pueda graficarlo aunque sea N/A
    if pd.notna(t_inicio_evento): ax.axvline(x=float(t_inicio_evento), color='purple', linestyle='--', label='Inicio Evento')
    if pd.notna(t_pico_energia): ax.axvline(x=float(t_pico_energia), color='orange', linestyle='--', label='Pico de Energía')
    if pd.notna(t_inicio_fusion): ax.axvline(x=float(t_inicio_fusion), color='brown', linestyle='--', label='Inicio Fusión')
    if pd.notna(t_fin_evento): ax.axvline(x=float(t_fin_evento), color='cyan', linestyle='--', label='Fin Evento')
        
    ax.set_title('Análisis de Señal de Fusión de Fusible')
    ax.set_xlabel('Tiempo (s)')
    ax.set_ylabel('Corriente (kA)')
    ax.legend(title='Componentes del Gráfico')
    ax.grid(True)
    return fig

# --- Función para Crear el Reporte de Excel ---
def create_excel_report(df_results, matplotlib_fig):
    """Genera un archivo de Excel con la tabla y el gráfico estático."""
    output_buffer = io.BytesIO()
    
    # Guardar el gráfico de Matplotlib en un buffer de imagen
    img_buffer = io.BytesIO()
    matplotlib_fig.savefig(img_buffer, format='png', bbox_inches='tight')
    img = OpenpyxlImage(img_buffer)

    # Crear el libro de trabajo de Excel
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        df_results.to_excel(writer, sheet_name='Resultados del Análisis', index=False, startrow=1)
        worksheet = writer.sheets['Resultados del Análisis']
        
        # Formatear la hoja
        worksheet.cell(row=1, column=1, value="Resultados del Análisis de Fusión")
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 25
        
        # Insertar la imagen
        worksheet.add_image(img, 'D2')

    return output_buffer.getvalue()

def create_point_analysis_matplotlib_figure(analyzed_df, t1, t2, actual_t1, actual_t2, actual_i1, actual_i2, selected_data):
    """Crea una figura estática con Matplotlib para el análisis entre puntos."""
    fig, ax = plt.subplots(figsize=(12, 6))
    
    # Plot original signal
    ax.plot(analyzed_df['time_s'], analyzed_df['current_kA'], 
            label='Señal Original i(t)', alpha=0.7, color='royalblue', linewidth=1.5)
    
    # Plot selected region with fill
    ax.fill_between(selected_data['time_s'], selected_data['current_kA'], 
                     alpha=0.3, color='orange', label='Región Seleccionada')
    ax.plot(selected_data['time_s'], selected_data['current_kA'], 
            color='orange', linewidth=3, label='Señal en Región')
    
    # Add vertical lines at selected points
    y_min, y_max = analyzed_df['current_kA'].min(), analyzed_df['current_kA'].max()
    ax.axvline(x=actual_t1, color='red', linestyle='--', linewidth=2, label=f'Punto 1: {actual_t1:.6f} s')
    ax.axvline(x=actual_t2, color='green', linestyle='--', linewidth=2, label=f'Punto 2: {actual_t2:.6f} s')
    
    # Add markers at the selected points
    ax.plot([actual_t1, actual_t2], [actual_i1, actual_i2], 
            'o', markersize=10, color='red', markeredgecolor='white', 
            markeredgewidth=2, label='Puntos Seleccionados', zorder=5)
    ax.text(actual_t1, actual_i1, ' P1', fontsize=10, color='red', 
            verticalalignment='bottom', fontweight='bold')
    ax.text(actual_t2, actual_i2, ' P2', fontsize=10, color='green', 
            verticalalignment='bottom', fontweight='bold')
    
    ax.set_title('Análisis de Señal Original entre Puntos Seleccionados', fontsize=14, fontweight='bold')
    ax.set_xlabel('Tiempo (s)', fontsize=12)
    ax.set_ylabel('Corriente (kA)', fontsize=12)
    ax.legend(loc='best', fontsize=10)
    ax.grid(True, alpha=0.3)
    
    return fig

def create_point_analysis_excel_report(delta_time, rms_current, num_points, t1, t2, matplotlib_fig):
    """Genera un archivo de Excel con los resultados del análisis entre puntos y el gráfico."""
    output_buffer = io.BytesIO()
    
    # Create results dataframe
    results_df = pd.DataFrame({
        'Parámetro': [
            'Tiempo Punto 1 (s)',
            'Tiempo Punto 2 (s)',
            'Diferencia de Tiempo Δt (s)',
            'Corriente RMS entre Puntos (kA)',
            'Número de Puntos de Datos'
        ],
        'Valor': [
            f'{t1:.6f}',
            f'{t2:.6f}',
            f'{delta_time:.6f}',
            f'{rms_current:.6f}',
            num_points
        ]
    })
    
    # Save matplotlib figure to image buffer
    img_buffer = io.BytesIO()
    matplotlib_fig.savefig(img_buffer, format='png', bbox_inches='tight', dpi=150)
    img = OpenpyxlImage(img_buffer)
    
    # Create Excel workbook
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='Análisis entre Puntos', index=False, startrow=1)
        worksheet = writer.sheets['Análisis entre Puntos']
        
        # Format the sheet
        worksheet.cell(row=1, column=1, value="Resultados del Análisis entre Puntos Seleccionados")
        worksheet.column_dimensions['A'].width = 45
        worksheet.column_dimensions['B'].width = 30
        
        # Insert the image
        worksheet.add_image(img, 'D2')
    
    return output_buffer.getvalue()

# --- Función de Análisis Principal ---
def analyze_signal(df, N, k):
    # 1. Preparación y Detección de Evento
    df.columns = ['time_ms', 'current_kA']
    df['time_s'] = df['time_ms'] / 1000.0
    noise_period_end_time_s = df['time_s'].quantile(0.3)
    noise_data = df[df['time_s'] <= noise_period_end_time_s]
    i_rms_base = np.sqrt(np.mean(noise_data['current_kA']**2))
    df['i_rms'] = np.sqrt(df['current_kA'].rolling(window=N, center=True).apply(lambda x: np.mean(x**2), raw=True).fillna(0))
    threshold = k * i_rms_base
    df['is_event'] = (df['i_rms'] > threshold).astype(int)
    event_start_indices = np.where(df['is_event'].diff() == 1)[0]
    t_inicio_evento = df['time_s'].iloc[event_start_indices[0]] if len(event_start_indices) > 0 else None
    event_end_indices = np.where(df['is_event'].diff() == -1)[0]
    t_fin_evento = None
    if t_inicio_evento is not None and len(event_end_indices) > 0:
        end_indices_after_start = event_end_indices[event_end_indices > event_start_indices[0]]
        if len(end_indices_after_start) > 0:
            t_fin_evento = df['time_s'].iloc[end_indices_after_start[0]]

    # 2. Identificación de Fases
    t_pico_energia, t_inicio_fusion = None, None
    I_rms_trans_inicial, I_rms_estable, I_rms_trans_final = None, None, None
    if t_inicio_evento and t_fin_evento:
        event_period_df = df[(df['time_s'] >= t_inicio_evento) & (df['time_s'] <= t_fin_evento)]
        if not event_period_df.empty:
            peak_rms_index = event_period_df['i_rms'].idxmax()
            t_pico_energia = df['time_s'].loc[peak_rms_index]
            peak_rms_value = df['i_rms'].loc[peak_rms_index]
            fusion_trigger_level = peak_rms_value * 0.85
            after_peak_df = event_period_df.loc[peak_rms_index:]
            fusion_candidates = after_peak_df[after_peak_df['i_rms'] < fusion_trigger_level]
            if not fusion_candidates.empty:
                t_inicio_fusion = df['time_s'].loc[fusion_candidates.index[0]]
            if t_pico_energia and t_inicio_fusion:
                trans_inicial_data = df[(df['time_s'] >= t_inicio_evento) & (df['time_s'] < t_pico_energia)]
                if not trans_inicial_data.empty: I_rms_trans_inicial = np.sqrt(np.mean(trans_inicial_data['current_kA']**2))
                estable_data = df[(df['time_s'] >= t_pico_energia) & (df['time_s'] < t_inicio_fusion)]
                if not estable_data.empty: I_rms_estable = np.sqrt(np.mean(estable_data['current_kA']**2))
                trans_final_data = df[(df['time_s'] >= t_inicio_fusion) & (df['time_s'] <= t_fin_evento)]
                if not trans_final_data.empty: I_rms_trans_final = np.sqrt(np.mean(trans_final_data['current_kA']**2))

    # 3. Crear DataFrame de Resultados
    results_df = pd.DataFrame({
        'Parámetro': ['Ruido RMS de Base (A)', 'Umbral Adaptativo (A)', 'Tiempo de Inicio del Evento (s)', 'Tiempo del Pico de Energía (s)', 'Tiempo de Inicio de Fusión (s)', 'Tiempo de Fin del Evento (s)', 'Corriente RMS Transitorio Inicial (A)', 'Corriente RMS Estado Estable (A)', 'Corriente RMS Transitorio Final (A)'],
        'Valor': [i_rms_base, threshold, t_inicio_evento, t_pico_energia, t_inicio_fusion, t_fin_evento, I_rms_trans_inicial, I_rms_estable, I_rms_trans_final]
    })
    results_df['Valor'] = results_df['Valor'].apply(lambda x: f'{x:.7f}' if isinstance(x, (int, float)) else 'N/A')
    
    # 4. Crear Gráfico Interactivo de Plotly
    plotly_fig = go.Figure()
    plotly_fig.add_trace(go.Scatter(x=df['time_s'], y=df['current_kA'], mode='lines', name='Señal Original i(t)', line=dict(color='royalblue'), hovertemplate='<b>Tiempo:</b> %{x:.6f} s<br><b>Corriente:</b> %{y:.6f} kA<extra></extra>'))
    plotly_fig.add_trace(go.Scatter(x=df['time_s'], y=df['i_rms'], mode='lines', name='RMS Móvil i_rms(t)', line=dict(color='red'), hoverinfo='skip'))
    y_range = [df['current_kA'].min(), df['current_kA'].max()]
    plotly_fig.add_trace(go.Scatter(x=[df['time_s'].iloc[0], df['time_s'].iloc[-1]], y=[threshold, threshold], mode='lines', name='Umbral', line=dict(color='green', dash='dash')))
    if t_inicio_evento: plotly_fig.add_trace(go.Scatter(x=[t_inicio_evento, t_inicio_evento], y=y_range, mode='lines', name='Inicio Evento', line=dict(color='purple', dash='dash')))
    if t_pico_energia: plotly_fig.add_trace(go.Scatter(x=[t_pico_energia, t_pico_energia], y=y_range, mode='lines', name='Pico de Energía', line=dict(color='orange', dash='dash')))
    if t_inicio_fusion: plotly_fig.add_trace(go.Scatter(x=[t_inicio_fusion, t_inicio_fusion], y=y_range, mode='lines', name='Inicio Fusión', line=dict(color='brown', dash='dash')))
    if t_fin_evento: plotly_fig.add_trace(go.Scatter(x=[t_fin_evento, t_fin_evento], y=y_range, mode='lines', name='Fin Evento', line=dict(color='cyan', dash='dash')))
    plotly_fig.update_layout(title='Análisis Interactivo de Señal de Fusión de Fusibles', xaxis_title='Tiempo (s)', yaxis_title='Corriente (kA)', legend_title='Componentes del Gráfico', hovermode='x unified')

    return plotly_fig, results_df, threshold, df

# --- Nueva función para crear un reporte completo con ambos análisis
def create_combined_excel_report(main_results_df, main_matplotlib_fig, point_analysis_data=None):
    """Genera un archivo de Excel completo con el análisis principal y el análisis entre puntos."""
    output_buffer = io.BytesIO()
    
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        # Sheet 1: Main Analysis
        main_results_df.to_excel(writer, sheet_name='Análisis Principal', index=False, startrow=1)
        worksheet1 = writer.sheets['Análisis Principal']
        worksheet1.cell(row=1, column=1, value="Resultados del Análisis de Fusión")
        worksheet1.column_dimensions['A'].width = 40
        worksheet1.column_dimensions['B'].width = 25
        
        # Add main analysis chart
        img_buffer1 = io.BytesIO()
        main_matplotlib_fig.savefig(img_buffer1, format='png', bbox_inches='tight', dpi=150)
        img1 = OpenpyxlImage(img_buffer1)
        worksheet1.add_image(img1, 'D2')
        
        # Sheet 2: Point Analysis (if data is provided)
        if point_analysis_data is not None:
            point_results_df = pd.DataFrame({
                'Parámetro': [
                    'Tiempo Punto 1 (s)',
                    'Tiempo Punto 2 (s)',
                    'Diferencia de Tiempo Δt (s)',
                    'Corriente RMS entre Puntos (kA)',
                    'Número de Puntos de Datos'
                ],
                'Valor': [
                    f"{point_analysis_data['t1']:.6f}",
                    f"{point_analysis_data['t2']:.6f}",
                    f"{point_analysis_data['delta_time']:.6f}",
                    f"{point_analysis_data['rms_current']:.6f}",
                    point_analysis_data['num_points']
                ]
            })
            
            point_results_df.to_excel(writer, sheet_name='Análisis entre Puntos', index=False, startrow=1)
            worksheet2 = writer.sheets['Análisis entre Puntos']
            worksheet2.cell(row=1, column=1, value="Resultados del Análisis entre Puntos Seleccionados")
            worksheet2.column_dimensions['A'].width = 45
            worksheet2.column_dimensions['B'].width = 30
            
            # Add point analysis chart
            img_buffer2 = io.BytesIO()
            point_analysis_data['matplotlib_fig'].savefig(img_buffer2, format='png', bbox_inches='tight', dpi=150)
            img2 = OpenpyxlImage(img_buffer2)
            worksheet2.add_image(img2, 'D2')
    
    return output_buffer.getvalue()

# --- Interfaz de la Aplicación Streamlit ---
st.header("Parámetros de Análisis")
uploaded_file = st.file_uploader("Elige un archivo CSV", type="csv")

param_col1, param_col2 = st.columns(2)
with param_col1:
    N = st.slider("Tamaño de Ventana (N)", 10, 200, 50, help="Número de muestras para calcular el RMS móvil")
with param_col2:
    k = st.slider("Factor de Umbral (k)", 0.1, 15.0, 5.0, 0.1, help="Multiplicador del RMS base para definir el umbral de detección")

st.markdown("---")

if uploaded_file is not None:
    try:
        df = pd.read_csv(uploaded_file)
        plotly_fig, results_df, threshold, analyzed_df = analyze_signal(df.copy(), N, k)
        
        st.header("Resultados del Análisis")
        
        matplotlib_fig = create_matplotlib_figure(analyzed_df, results_df, threshold)
        
        col1, col2 = st.columns([1, 1])
        with col1:
            st.dataframe(results_df)
        
        with col2:
            st.plotly_chart(plotly_fig, use_container_width=True)
        
        st.header("Análisis de Señal Original")
        st.write("Usa los controles deslizantes para seleccionar dos puntos en la señal y calcular el tiempo transcurrido y el RMS entre ellos.")
        
        # Create two columns for slider controls
        col_s1, col_s2 = st.columns(2)
        
        min_time = float(analyzed_df['time_s'].min())
        max_time = float(analyzed_df['time_s'].max())
        
        slider_step = 0.000001  # 1 microsecond for slider (easier to drag)
        button_step = 0.0000001  # 0.1 microsecond for buttons (fine precision)
        
        if 'time_point1' not in st.session_state:
            st.session_state.time_point1 = min_time
        if 'time_point2' not in st.session_state:
            st.session_state.time_point2 = max_time
        
        with col_s1:
            st.subheader("Punto 1")
            
            btn_col1, slider_col1, btn_col2 = st.columns([1, 8, 1])
            
            with btn_col1:
                st.write("")  # Spacing
                st.write("")  # Spacing
                if st.button("➖", key="dec_p1", help="Retroceder 0.1 microsegundos"):
                    st.session_state.time_point1 = max(min_time, st.session_state.time_point1 - button_step)
                    st.rerun()
            
            with slider_col1:
                time_point1 = st.slider(
                    "Tiempo (s)", 
                    min_value=min_time, 
                    max_value=max_time,
                    value=st.session_state.time_point1,
                    step=slider_step,
                    format="%.7f",
                    key="slider_p1",
                    help="Arrastra para ajustes rápidos (1 µs) o usa botones +/- para precisión fina (0.1 µs)"
                )
                st.session_state.time_point1 = time_point1
            
            with btn_col2:
                st.write("")  # Spacing
                st.write("")  # Spacing
                if st.button("➕", key="inc_p1", help="Avanzar 0.1 microsegundos"):
                    st.session_state.time_point1 = min(max_time, st.session_state.time_point1 + button_step)
                    st.rerun()
            
            st.caption(f"Tiempo seleccionado: {st.session_state.time_point1:.7f} s")
        
        with col_s2:
            st.subheader("Punto 2")
            
            btn_col3, slider_col2, btn_col4 = st.columns([1, 8, 1])
            
            with btn_col3:
                st.write("")  # Spacing
                st.write("")  # Spacing
                if st.button("➖", key="dec_p2", help="Retroceder 0.1 microsegundos"):
                    st.session_state.time_point2 = max(min_time, st.session_state.time_point2 - button_step)
                    st.rerun()
            
            with slider_col2:
                time_point2 = st.slider(
                    "Tiempo (s)", 
                    min_value=min_time, 
                    max_value=max_time,
                    value=st.session_state.time_point2,
                    step=slider_step,
                    format="%.7f",
                    key="slider_p2",
                    help="Arrastra para ajustes rápidos (1 µs) o usa botones +/- para precisión fina (0.1 µs)"
                )
                st.session_state.time_point2 = time_point2
            
            with btn_col4:
                st.write("")  # Spacing
                st.write("")  # Spacing
                if st.button("➕", key="inc_p2", help="Avanzar 0.1 microsegundos"):
                    st.session_state.time_point2 = min(max_time, st.session_state.time_point2 + button_step)
                    st.rerun()
            
            st.caption(f"Tiempo seleccionado: {st.session_state.time_point2:.7f} s")
        
        t1 = st.session_state.time_point1
        t2 = st.session_state.time_point2
        
        # Ensure t1 < t2
        if t1 > t2:
            t1, t2 = t2, t1
        
        # Calculate time difference
        delta_time = t2 - t1
        
        # Filter data between the two points
        selected_data = analyzed_df[(analyzed_df['time_s'] >= t1) & (analyzed_df['time_s'] <= t2)]
        
        point_analysis_data = None
        
        # Calculate RMS current between the two points
        if not selected_data.empty:
            rms_current = np.sqrt(np.mean(selected_data['current_kA']**2))
            
            # Find closest actual data points for visualization
            idx1 = (analyzed_df['time_s'] - t1).abs().idxmin()
            idx2 = (analyzed_df['time_s'] - t2).abs().idxmin()
            actual_t1 = analyzed_df.loc[idx1, 'time_s']
            actual_t2 = analyzed_df.loc[idx2, 'time_s']
            actual_i1 = analyzed_df.loc[idx1, 'current_kA']
            actual_i2 = analyzed_df.loc[idx2, 'current_kA']
            
            # Display metrics prominently
            st.subheader("Resultados del Análisis entre Puntos")
            metric_col1, metric_col2, metric_col3 = st.columns(3)
            with metric_col1:
                st.metric("Diferencia de Tiempo (Δt)", f"{delta_time:.7f} s")
            with metric_col2:
                st.metric("Corriente RMS", f"{rms_current:.6f} kA")
            with metric_col3:
                st.metric("Puntos de Datos", f"{len(selected_data)}")
            
            point_analysis_fig = create_point_analysis_matplotlib_figure(
                analyzed_df, t1, t2, actual_t1, actual_t2, actual_i1, actual_i2, selected_data
            )
            
            point_analysis_data = {
                't1': actual_t1,
                't2': actual_t2,
                'delta_time': delta_time,
                'rms_current': rms_current,
                'num_points': len(selected_data),
                'matplotlib_fig': point_analysis_fig
            }
            
            # Create interactive plot with selected points
            original_signal_fig = go.Figure()
            
            # Add original signal
            original_signal_fig.add_trace(
                go.Scatter(
                    x=analyzed_df['time_s'], 
                    y=analyzed_df['current_kA'], 
                    mode='lines', 
                    name='Señal Original i(t)', 
                    line=dict(color='royalblue', width=2),
                    hovertemplate='<b>Tiempo:</b> %{x:.6f} s<br><b>Corriente:</b> %{y:.6f} kA<extra></extra>'
                )
            )
            
            # Add highlighted region between points with fill
            original_signal_fig.add_trace(
                go.Scatter(
                    x=selected_data['time_s'],
                    y=selected_data['current_kA'],
                    mode='lines',
                    name='Región Seleccionada',
                    line=dict(color='orange', width=4),
                    fill='tozeroy',
                    fillcolor='rgba(255, 165, 0, 0.2)',
                    hovertemplate='<b>Tiempo:</b> %{x:.6f} s<br><b>Corriente:</b> %{y:.6f} kA<extra></extra>'
                )
            )
            
            # Add vertical lines at selected points
            y_range = [analyzed_df['current_kA'].min(), analyzed_df['current_kA'].max()]
            original_signal_fig.add_trace(
                go.Scatter(
                    x=[actual_t1, actual_t1],
                    y=y_range,
                    mode='lines',
                    name=f'Punto 1: {actual_t1:.6f} s',
                    line=dict(color='red', width=3, dash='dash'),
                    showlegend=True
                )
            )
            original_signal_fig.add_trace(
                go.Scatter(
                    x=[actual_t2, actual_t2],
                    y=y_range,
                    mode='lines',
                    name=f'Punto 2: {actual_t2:.6f} s',
                    line=dict(color='green', width=3, dash='dash'),
                    showlegend=True
                )
            )
            
            # Add markers at the selected points
            original_signal_fig.add_trace(
                go.Scatter(
                    x=[actual_t1, actual_t2],
                    y=[actual_i1, actual_i2],
                    mode='markers+text',
                    name='Puntos Seleccionados',
                    marker=dict(size=12, color=['red', 'green'], symbol='circle', line=dict(width=2, color='white')),
                    text=['P1', 'P2'],
                    textposition='top center',
                    textfont=dict(size=12, color='white'),
                    hovertemplate='<b>Punto:</b> %{text}<br><b>Tiempo:</b> %{x:.6f} s<br><b>Corriente:</b> %{y:.6f} kA<extra></extra>'
                )
            )
            
            original_signal_fig.update_layout(
                title='Señal Original de Corriente vs Tiempo - Análisis Interactivo entre Puntos',
                xaxis_title='Tiempo (s)',
                yaxis_title='Corriente (kA)',
                hovermode='x unified',
                showlegend=True,
                height=600,
                legend=dict(
                    orientation="v",
                    yanchor="top",
                    y=0.99,
                    xanchor="right",
                    x=0.99
                )
            )
            
            st.plotly_chart(original_signal_fig, use_container_width=True)
        else:
            st.warning("No hay datos disponibles entre los puntos seleccionados.")
        
        st.markdown("---")
        st.header("Descargar Reporte Completo")
        st.write("Descarga un reporte en Excel con todos los análisis realizados.")
        
        combined_excel_data = create_combined_excel_report(results_df, matplotlib_fig, point_analysis_data)
        
        st.download_button(
            label="Descargar Reporte Completo en Excel",
            data=combined_excel_data,
            file_name="reporte_completo_analisis_fusion.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Descarga un reporte completo con el análisis principal y el análisis entre puntos seleccionados",
            use_container_width=True
        )

    except Exception as e:
        st.error(f"Ocurrió un error durante el análisis: {e}")
else:
    st.info("Esperando a que se suba un archivo CSV.")
